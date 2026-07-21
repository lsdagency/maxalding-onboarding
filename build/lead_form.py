"""
Meta Lead Form (Instant Form) copy generator.

Builds a branded LEAD FORM tab and, optionally, a combined workbook that places
the lead form copy alongside the AD COPY tab (reusing creative_plan._build_ad_copy
so the ad copy half is byte-identical to the onboarding Creative Plan).

Branding and styling come from build/template.py, so every tab matches the rest
of the Maxalding deliverables: Helvetica Neue, black section headers with white
bold text, white #434343 body, wrap on, top-aligned, frozen top row, live =LEN()
formulas on the length-limited fields.

Single source of truth for the rules: the meta-lead-form skill.

Usage:
    # combined ad copy + lead form (default)
    python -m build.lead_form "<data>.json" --out "<dir>"
    # lead form only
    python -m build.lead_form "<data>.json" --out "<dir>" --lead-form-only

Input JSON shape:
{
  "client_business_name": "FIT Republik",
  "ad_copy": { ... same shape as build.ad_copy ... },   # optional, for the combined file
  "lead_form": {
    "sections": [
      {"name": "FORM SETUP", "rows": [
        {"field": "Form name", "copy": "..."},
        {"field": "Form type", "copy": "..."}
      ]},
      {"name": "INTRO", "rows": [
        {"field": "Intro headline", "copy": "...", "limit": 60},
        {"field": "Intro description", "copy": "..."}
      ]},
      ...
    ]
  }
}
A row with a "limit" gets a live =LEN() in the CHARS column and is checked
against that Meta character limit at build time.
"""
from __future__ import annotations

import argparse
import json
import os

import openpyxl

from . import template as T
from .creative_plan import _build_ad_copy

# Meta instant form character limits (used for validation and the CHARS column).
LIMITS = {
    "intro headline": 60,
    "completion headline": 60,
    "cta button text": 30,
}


def _section_header(ws, row, text, span=3):
    """Black section header row spanning the tab width."""
    for c in range(1, span + 1):
        cell = ws.cell(row=row, column=c, value=text if c == 1 else None)
        T.style_header_cell(cell)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    ws.row_dimensions[row].height = T.XLSX_MIN_ROW_HEIGHT


def _build_lead_form(wb, data):
    ws = wb.create_sheet("LEAD FORM")
    T.set_column_widths(ws, {"A": 26, "B": 95, "C": 10})

    for i, name in enumerate(["FIELD", "COPY", "CHARS"], start=1):
        T.style_header_cell(ws.cell(row=1, column=i, value=name))
    ws.row_dimensions[1].height = T.XLSX_MIN_ROW_HEIGHT
    ws.freeze_panes = "A2"

    lf = data.get("lead_form", {})
    row = 2
    for section in lf.get("sections", []):
        _section_header(ws, row, section.get("name", ""))
        row += 1
        for r in section.get("rows", []):
            field = r.get("field", "")
            copy = r.get("copy", "")
            label = ws.cell(row=row, column=1, value=field)
            T.style_body_cell(label)
            label.font = T.xlsx_font(bold=True)
            T.style_body_cell(ws.cell(row=row, column=2, value=copy))
            ccell = ws.cell(row=row, column=3)
            T.style_body_cell(ccell)
            if r.get("limit"):
                ccell.value = f"=LEN(B{row})"
            ws.row_dimensions[row].height = T.XLSX_MIN_ROW_HEIGHT
            row += 1
    return ws


def validate(data) -> list[str]:
    problems: list[str] = []
    lf = data.get("lead_form", {})
    sections = lf.get("sections", [])
    if not sections:
        problems.append("lead_form.sections is empty")
    for section in sections:
        for r in section.get("rows", []):
            limit = r.get("limit")
            copy = r.get("copy", "")
            if limit and len(copy) > limit:
                problems.append(f"'{r.get('field')}' over {limit} ({len(copy)}): {copy}")
    return problems


def build(data, out_dir, lead_form_only=False) -> str:
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    if not lead_form_only and data.get("ad_copy"):
        _build_ad_copy(wb, data)
    _build_lead_form(wb, data)
    wb.remove(default_sheet)
    deliverable = "Meta Lead Form Copy" if lead_form_only else "Meta Ad & Lead Form Copy"
    filename = T.deliverable_filename(data["client_business_name"], deliverable,
                                      data.get("campaign"))
    out_path = os.path.join(out_dir, filename)
    wb.save(out_path)
    return out_path


def main():
    ap = argparse.ArgumentParser(description="Build the Meta Lead Form (and Ad) copy XLSX.")
    ap.add_argument("data", help="Path to the data JSON.")
    ap.add_argument("--out", required=True, help="Output directory.")
    ap.add_argument("--lead-form-only", action="store_true",
                    help="Build only the LEAD FORM tab, not the combined ad + lead form file.")
    ap.add_argument("--skip-validate", action="store_true")
    args = ap.parse_args()

    with open(args.data) as f:
        data = json.load(f)

    if not args.skip_validate:
        problems = validate(data)
        if problems:
            print("Validation failed:")
            for p in problems:
                print("  -", p)
            raise SystemExit(1)

    os.makedirs(args.out, exist_ok=True)
    path = build(data, args.out, lead_form_only=args.lead_form_only)
    print("Saved:", path)


if __name__ == "__main__":
    main()
