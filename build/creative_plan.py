"""
Creative Plan XLSX generator (System Prompt section 8).

Three tabs:
  1. BRAND PACK       label/value with black section headers, NO KPIs section
  2. CREATIVE TRACKER 13 rows (5 statics then 8 videos), dropdowns, live formulas
  3. AD COPY          distinct concepts, 5 post variations each, shared headlines
                      and one description, all with live =LEN() formulas

The generator takes a structured `data` dict (the campaign content the agents
produce) and lays it out to spec. See evals/cases/sample_client.json for the
exact shape. Styling comes from build/template.py.
"""

from __future__ import annotations

from . import template as T

RESOLUTION_STRING = "1080 x 1350 (4:5) / 1080 x 1920 (9:16)"

STATUS_OPTIONS = ["Briefed", "Filmed", "In Edit", "Signed Off", "Live"]
PLATFORM_OPTIONS = ["FACEBOOK + INSTAGRAM", "FACEBOOK", "INSTAGRAM", "TIKTOK"]
DEFAULT_PLATFORM = "FACEBOOK + INSTAGRAM"
DEFAULT_STATUS = "Briefed"

# Canonical 13-row tracker template (System Prompt section 8 table).
CANONICAL_TRACKER = [
    ("STATIC", "Problem > Solution"),
    ("STATIC", "Incentive"),
    ("STATIC", "Benefit"),
    ("STATIC", "Social Proof (Review)"),
    ("STATIC", "Social Proof (Before & After)"),
    ("VIDEO", "Problem > Solution"),
    ("VIDEO", "Social Proof"),
    ("VIDEO", "Pattern Interrupt"),
    ("VIDEO", "Storytelling"),
    ("VIDEO", "Benefit"),
    ("VIDEO", "Incentive"),
    ("VIDEO", "Social Proof"),
    ("VIDEO", "Audience Addresser"),
]

TRACKER_COLUMNS = [
    "#", "DATE", "FORMAT", "RESOLUTION", "STATUS", "PLATFORM", "FILE TYPE",
    "AUDIENCE", "CONCEPT", "HOOK", "BEATS", "DETAILS", "EDITING DIRECTION",
    "FILMING DIRECTION", "SCRIPT", "FINAL ASSET",
]

TRACKER_WIDTHS = {
    "A": 5, "B": 12, "C": 10, "D": 28, "E": 12, "F": 22, "G": 10, "H": 30,
    "I": 18, "J": 38, "K": 34, "L": 40, "M": 34, "N": 34, "O": 30, "P": 16,
}


def _write_row(ws, row, values):
    for i, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=value)
        T.style_body_cell(cell)
    return row


# ---------------------------------------------------------------------------
# Tab 1: BRAND PACK
# ---------------------------------------------------------------------------
def _build_brand_pack(wb, data):
    ws = wb.active
    ws.title = "BRAND PACK"
    T.set_column_widths(ws, {"A": 34, "B": 90})

    bp = data["brand_pack"]
    # Section order is fixed. No KPIs section (Feedback 2026-06-18).
    sections = [
        ("CHANNELS", bp.get("channels", [])),
        ("AUDIENCE", bp.get("audience", [])),
        ("TONE OF VOICE", bp.get("tone_of_voice", [])),
        ("OFFER", bp.get("offer", [])),
        ("BRAND POSITIONING", bp.get("brand_positioning", [])),
    ]

    row = 1
    for title, rows in sections:
        # Black section header spanning both columns.
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=title)
        T.style_header_cell(cell)
        ws.row_dimensions[row].height = T.XLSX_MIN_ROW_HEIGHT
        row += 1
        for label, value in rows:
            lab = ws.cell(row=row, column=1, value=label)
            T.style_body_cell(lab)
            lab.font = T.xlsx_font(bold=True)
            val = ws.cell(row=row, column=2, value=value)
            T.style_body_cell(val)
            ws.row_dimensions[row].height = T.XLSX_MIN_ROW_HEIGHT
            row += 1
    return ws


# ---------------------------------------------------------------------------
# Tab 2: CREATIVE TRACKER
# ---------------------------------------------------------------------------
def _build_tracker(wb, data):
    ws = wb.create_sheet("CREATIVE TRACKER")
    T.set_column_widths(ws, TRACKER_WIDTHS)

    # Header row.
    for i, name in enumerate(TRACKER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=i, value=name)
        T.style_header_cell(cell)
    ws.row_dimensions[1].height = T.XLSX_MIN_ROW_HEIGHT
    ws.freeze_panes = "A2"

    audience_line = data.get("audience_line", "")
    date_value = data.get("date", "")
    rows = data.get("tracker", [])
    video_scripts_file = data.get("video_scripts_filename", "")

    for idx, (fmt, concept) in enumerate(CANONICAL_TRACKER):
        r = idx + 2
        content = rows[idx] if idx < len(rows) else {}
        is_static = fmt == "STATIC"
        script_value = "" if is_static else (content.get("script") or video_scripts_file)
        values = [
            idx + 1,                                   # A #
            date_value,                                # B DATE (blank by default, no date stamped)
            fmt,                                       # C FORMAT
            RESOLUTION_STRING,                         # D RESOLUTION
            content.get("status", DEFAULT_STATUS),     # E STATUS
            content.get("platform", DEFAULT_PLATFORM), # F PLATFORM
            None,                                      # G FILE TYPE (formula)
            audience_line,                             # H AUDIENCE (identical)
            concept,                                   # I CONCEPT
            content.get("hook", ""),                   # J HOOK
            content.get("beats", ""),                  # K BEATS
            content.get("details", ""),                # L DETAILS
            content.get("editing", ""),                # M EDITING DIRECTION
            content.get("filming", ""),                # N FILMING DIRECTION
            script_value,                              # O SCRIPT
            content.get("final_asset", ""),            # P FINAL ASSET
        ]
        _write_row(ws, r, values)
        # FILE TYPE is always a formula referencing the FORMAT column (C).
        ws.cell(row=r, column=7, value=f'=IF(C{r}="STATIC","JPG","MP4")')
        T.style_body_cell(ws.cell(row=r, column=7))
        ws.row_dimensions[r].height = T.XLSX_TRACKER_ROW_HEIGHT

    last = len(CANONICAL_TRACKER) + 1
    # Dropdowns on STATUS (E) and PLATFORM (F).
    T.add_dropdown(ws, f"E2:E{last}", STATUS_OPTIONS)
    T.add_dropdown(ws, f"F2:F{last}", PLATFORM_OPTIONS)
    return ws


# ---------------------------------------------------------------------------
# Tab 3: AD COPY
# ---------------------------------------------------------------------------
def _build_ad_copy(wb, data):
    ws = wb.create_sheet("AD COPY")
    T.set_column_widths(ws, {"A": 5, "B": 22, "C": 90, "D": 10})

    # Header row: # | CONCEPT | POST COPY | CHARS. No STATUS, no CTA columns.
    for i, name in enumerate(["#", "CONCEPT", "POST COPY", "CHARS"], start=1):
        cell = ws.cell(row=1, column=i, value=name)
        T.style_header_cell(cell)
    ws.row_dimensions[1].height = T.XLSX_MIN_ROW_HEIGHT
    ws.freeze_panes = "A2"

    ad = data.get("ad_copy", {})
    row = 2
    n = 1
    for block in ad.get("concepts", []):
        concept = block.get("concept", "")
        for post in block.get("posts", []):
            ws.cell(row=row, column=1, value=n)
            ws.cell(row=row, column=2, value=concept)
            ws.cell(row=row, column=3, value=post)
            ws.cell(row=row, column=4, value=f"=LEN(C{row})")
            for c in range(1, 5):
                T.style_body_cell(ws.cell(row=row, column=c))
            ws.row_dimensions[row].height = T.XLSX_MIN_ROW_HEIGHT
            row += 1
            n += 1

    # Blank spacer row.
    row += 1

    # Shared HEADLINES block (label cell drives the validator's headline tag).
    hcell = ws.cell(row=row, column=2, value="HEADLINES")
    T.style_body_cell(hcell)
    hcell.font = T.xlsx_font(bold=True)
    row += 1
    for headline in ad.get("headlines", []):
        ws.cell(row=row, column=2, value=headline)
        ws.cell(row=row, column=4, value=f"=LEN(B{row})")
        T.style_body_cell(ws.cell(row=row, column=2))
        T.style_body_cell(ws.cell(row=row, column=4))
        ws.row_dimensions[row].height = T.XLSX_MIN_ROW_HEIGHT
        row += 1

    row += 1

    # One DESCRIPTION.
    dcell = ws.cell(row=row, column=2, value="DESCRIPTION")
    T.style_body_cell(dcell)
    dcell.font = T.xlsx_font(bold=True)
    row += 1
    description = ad.get("description", "")
    ws.cell(row=row, column=2, value=description)
    ws.cell(row=row, column=4, value=f"=LEN(B{row})")
    T.style_body_cell(ws.cell(row=row, column=2))
    T.style_body_cell(ws.cell(row=row, column=4))
    return ws


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def build_creative_plan(data, out_dir):
    import os
    import openpyxl

    wb = openpyxl.Workbook()
    _build_brand_pack(wb, data)
    _build_tracker(wb, data)
    _build_ad_copy(wb, data)

    filename = T.deliverable_filename(data["client_business_name"], "Creative Plan")
    out_path = os.path.join(out_dir, filename)
    wb.save(out_path)
    return out_path
